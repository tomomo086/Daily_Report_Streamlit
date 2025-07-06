from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from datetime import datetime
import io
from models import PatrolData
from utils.time_utils import PatrolTimeGenerator
from .cell_definitions import CellDefinitionManager

class ExcelWriter:
    def __init__(self):
        self.time_generator = PatrolTimeGenerator()
        self.cell_manager = CellDefinitionManager()
    
    def write_report(self, patrol_data: PatrolData):
        """日報をExcelファイルに書き込む（テンプレート不要）"""
        wb = Workbook()
        
        # シート名の取得
        today = datetime.today()
        sheet_name = f"{today.month}.{today.day}"
        
        ws = wb.active
        ws.title = sheet_name
        
        # 日報テンプレートの基本構造を作成
        self._create_template_structure(ws)
        
        # 日付を自動入力
        self._write_date(ws, today)
        
        # 基本情報の書き込み
        self._write_basic_info(ws, patrol_data)
        
        # 巡回記録の書き込み
        self._write_patrol_records(ws, patrol_data)
        
        # その他の時間記録
        self._write_other_records(ws, patrol_data)
        
        # フォントサイズの設定
        self._set_font_sizes(wb)
        
        # バイト配列として返す
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_template_structure(self, ws):
        """日報テンプレートの基本構造を作成"""
        # タイトル
        ws['A1'] = '日報'
        ws['A1'].font = Font(size=16, bold=True)
        
        # 日付欄
        ws['A2'] = '日付:'
        ws['B2'] = ''  # 日付が入る場所
        
        # 基本情報のヘッダー
        ws['A4'] = '天気'
        ws['I4'] = ''  # 天気が入る場所
        
        # 担当者情報
        ws['A6'] = '4ポスト担当'
        ws['F6'] = ''  # 4ポスト担当者が入る場所
        ws['A7'] = '5ポスト担当'  
        ws['F7'] = ''  # 5ポスト担当者が入る場所
        
        # 設備担当者
        ws['A5'] = '設備担当者'
        ws['J5'] = ''  # 設備担当者が入る場所
        ws['J6'] = ''  # 設備担当者が入る場所（2箇所目）
        
        # 勤務時間
        ws['A10'] = '勤務時間'
        ws['K4'] = ''  # 勤務時間1
        ws['L4'] = ''  # 勤務時間2
        
        # 巡回記録のヘッダー
        ws['A15'] = '巡回記録'
        ws['A17'] = '4ポスト巡回'
        ws['A26'] = '5ポスト巡回'
        ws['A32'] = 'その他時間記録'
        
        # セルの基本スタイル設定
        self._apply_basic_styling(ws)
    
    def _write_date(self, ws, today):
        """日付を自動入力"""
        date_str = today.strftime('%Y年%m月%d日')
        ws['B2'] = date_str
        ws['B2'].font = Font(size=12)
    
    def _apply_basic_styling(self, ws):
        """基本的なスタイリングを適用"""
        # ヘッダー行のスタイル
        header_font = Font(bold=True, size=10)
        for row in [1, 4, 6, 7, 15, 17, 26, 32]:
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.font = header_font
    
    def _safe_set_cell_value(self, ws, cell_address, value):
        """結合セルかどうかをチェックしてから値を設定"""
        try:
            cell = ws[cell_address]
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = merged_range.start_cell
                    ws[top_left.coordinate] = value
                    return
            cell.value = value
        except Exception as e:
            # エラーが発生しても処理を継続
            try:
                ws[cell_address] = value
            except Exception as fallback_error:
                pass
    
    def _set_time(self, ws, cell, time_str):
        """時間をセルに設定（0埋め除去対応）"""
        if time_str:
            try:
                if isinstance(time_str, str):
                    # 0埋め除去
                    time_str = time_str.lstrip("0") if time_str != "-" else time_str
                    # 文字列として書き込む
                    self._safe_set_cell_value(ws, cell, time_str)
                else:
                    self._safe_set_cell_value(ws, cell, str(time_str).lstrip("0"))
            except Exception as time_error:
                self._safe_set_cell_value(ws, cell, time_str)
    
    def _write_basic_info(self, ws, patrol_data: PatrolData):
        """基本情報を書き込む"""
        # セル定義管理を使用して安全に書き込み
        self._safe_set_cell_value(ws, 'I4', patrol_data.weather)
        self._safe_set_cell_value(ws, 'F6', patrol_data.post4)
        self._safe_set_cell_value(ws, 'F7', patrol_data.post5)
        # 設備担当者は登録された通りに出力（そのまま）
        self._safe_set_cell_value(ws, 'J5', patrol_data.supervisor)
        self._safe_set_cell_value(ws, 'J6', patrol_data.supervisor)

        # 勤務区分による分岐
        if getattr(patrol_data, 'work_type', '通常') == '早出':
            # 早出
            self._safe_set_cell_value(ws, 'K4', '7:30～23:00')
            self._safe_set_cell_value(ws, 'L4', '7:30～23:00')
            # C10, D10は0埋めなしの文字列で書き込む
            self._safe_set_cell_value(ws, 'C10', '7:30')
            self._safe_set_cell_value(ws, 'D10', '7:30')
            self._safe_set_cell_value(ws, 'E10', patrol_data.post4)
            self._safe_set_cell_value(ws, 'F10', patrol_data.post4)
            self._set_time(ws, 'C11', '23:00')
            self._safe_set_cell_value(ws, 'E11', patrol_data.post4)
        elif getattr(patrol_data, 'work_type', '通常') == '残業':
            # 残業
            self._safe_set_cell_value(ws, 'K4', '8:00～24:00')
            self._safe_set_cell_value(ws, 'L4', '8:00～24:00')
            # C10は従来通り（例として）
            self._set_time(ws, 'C10', '8:00')
            self._safe_set_cell_value(ws, 'E10', patrol_data.post1)
            self._set_time(ws, 'C11', '24:00')
            self._set_time(ws, 'D11', '24:00')
            self._safe_set_cell_value(ws, 'E11', patrol_data.post4)
        else:
            # 通常
            self._set_time(ws, 'C10', '8:00')
            self._safe_set_cell_value(ws, 'E10', patrol_data.post1)
            self._set_time(ws, 'C11', '23:00')
            self._safe_set_cell_value(ws, 'E11', patrol_data.post4)
    
    def _write_patrol_records(self, ws, patrol_data: PatrolData):
        """巡回記録を書き込む"""
        # 4ポストの巡回
        records = self.time_generator.generate_4post_times(
            patrol_data.patrol_start,
            patrol_data.large_theater_used,
            patrol_data.medium_theater_used,
            patrol_data.small_theater_used
        )
        
        for cell_start, cell_end, record in records:
            if record.start_time != "-":
                self._set_time(ws, cell_start, record.start_time)
            else:
                self._safe_set_cell_value(ws, cell_start, "-")
                
            if record.end_time != "-":
                self._set_time(ws, cell_end, record.end_time)
            else:
                self._safe_set_cell_value(ws, cell_end, "-")
            
            # コメントの書き込み
            comment_cell = cell_start.replace('C', 'F').replace('E', 'F')
            self._safe_set_cell_value(ws, comment_cell, record.comment)
        
        # 5ポストの巡回
        records_5post = self.time_generator.generate_5post_times(
            patrol_data.large_theater_used,
            patrol_data.medium_theater_used,
            patrol_data.small_theater_used
        )
        
        for cell_start, cell_end, record in records_5post:
            self._set_time(ws, cell_start, record.start_time)
            self._set_time(ws, cell_end, record.end_time)
            comment_cell = cell_start.replace('C', 'F')
            self._safe_set_cell_value(ws, comment_cell, record.comment)
    
    def _write_other_records(self, ws, patrol_data: PatrolData):
        """その他の時間記録を書き込む"""
        other_times = self.time_generator.generate_other_times()
        
        for cell, time, name_attr in [
            ('E32', other_times['morning_4post'], 'post5_lastname'),
            ('E34', other_times['morning_5post'], 'post5_lastname'),
            ('E36', other_times['morning_1post'], 'post1_lastname'),
            ('E38', other_times['morning_4post_2'], 'post4_lastname')
        ]:
            self._safe_set_cell_value(ws, cell, time.lstrip("0"))
            name_cell = cell.replace('E', 'G')
            self._safe_set_cell_value(ws, name_cell, 
                                    getattr(patrol_data, name_attr))
        
        # 夜の記録
        if getattr(patrol_data, 'work_type', '通常') == '残業':
            # 残業の場合はH34とI34に23:50を入力
            self._set_time(ws, 'H34', "23:50")
            self._safe_set_cell_value(ws, 'I34', "23:50")
        else:
            self._set_time(ws, 'H34', "22:50")
        self._safe_set_cell_value(ws, 'J34', patrol_data.post5_lastname)
        
        self._set_time(ws, 'H38', other_times['night_4post'])
        self._safe_set_cell_value(ws, 'J38', patrol_data.post4_lastname)
        
        self._set_time(ws, 'E41', other_times['patrol_4post'])
        self._safe_set_cell_value(ws, 'G41', patrol_data.post4_lastname)
        self._safe_set_cell_value(ws, 'F41', "22:50")
        
        self._set_time(ws, 'H41', other_times['patrol_4post_end'])
        self._safe_set_cell_value(ws, 'J41', patrol_data.post4_lastname)
    
    def _set_font_sizes(self, wb):
        """特定セルのフォントサイズを設定"""
        for ws_item in wb.worksheets:
            for cell in ['I5', 'I6', 'K5', 'K6']:
                try:
                    ws_item[cell].font = Font(size=8)
                except Exception as font_error:
                    pass

